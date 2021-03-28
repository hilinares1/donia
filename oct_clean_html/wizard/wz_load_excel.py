import string

from odoo import api, fields, models, exceptions, _
import logging
import base64
import openpyxl
import tempfile
import string
import random
import re

_logger = logging.getLogger()


class LoadExcelFile(models.TransientModel):
    _name = 'wizard.load_html_file'
    _description = 'Model to import and clean data'

    serial_file = fields.Binary(string="Load File")
    col_html = fields.Integer(string='Html column')

    def generate_random_name(self):
        letters = string.ascii_lowercase
        return ''.join(random.choice(letters) for i in range(10))

    def load_html_file(self):

        if not self.col_attribute or not self.col_attribute_value:
            raise exceptions.Warning(_("The fields [html column] is required"))
        else:
            if not self.serial_file:
                raise exceptions.Warning(_("You must select a file to modify"))
            else:

                try:
                    final_name = "modified_file_" + self.generate_random_name()

                    with tempfile.NamedTemporaryFile(prefix=final_name, suffix='.xlsx', delete=False) as tmp:
                        tmp.write(base64.decodebytes(self.serial_file))

                    xfile = openpyxl.load_workbook(tmp.name)
                    sheet = xfile.worksheets[0]

                    row_count = sheet.max_row

                    row_with_value = 0

                    col_html = self.col_html

                    for row in range(1, row_count + 1):
                        if sheet.cell(row, col_html).value == None:
                            continue
                        else:

                            cleanr = re.compile('<.*?>')
                            cleantext = re.sub(cleanr, '', sheet.cell(row, col_html).value)
                            cleantext = cleantext.replace('   ', ' ')
                            cleantext = cleantext.replace('  ', ' ')
                            cleantext = cleantext.replace(
                                ' ENVÍO: Recibe el producto en tu establecimiento en tan solo 24/48 horas gracias a nuestro servicio de transporte urgente. ¿Tienes cualquier duda acerca del producto? Puedes ponerte en contacto con nosotros a través del email (info@vidalhome.es) o en el teléfono 968 884 620. ',
                                '')

                            sheet.cell(row, col_html).value = cleantext

                            _logger.info("====== VALORES HTML ======= %r " % cleantext)

                    xfile.save(tmp.name)

                    with open(tmp.name, 'rb') as r:
                        file = base64.b64encode(r.read())

                    att_vals = {
                        'name': final_name + '.xlsx',
                        'type': 'binary',
                        'datas': file,
                    }

                    attachment_id = self.env['ir.attachment'].create(att_vals)
                    self.env.cr.commit()

                    return {
                        'type': 'ir.actions.act_url',
                        'url': '/web/content/{}?download=true'.format(attachment_id.id, ),
                        'target': 'self',
                    }
                except Exception:
                    raise exceptions.Warning(_("The selected file seems to be not valid"))

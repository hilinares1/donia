import string

from odoo import api, fields, models, exceptions, _
import logging
import base64
import openpyxl
import tempfile
import string
import random

_logger = logging.getLogger()


class LoadExcelFile(models.TransientModel):
    _name = 'wizard.load_file'
    _description = 'Model to import and clean data'

    serial_file = fields.Binary(string="Load File")
    col_attribute = fields.Integer(string='Attribute column')
    col_attribute_value = fields.Integer(string='Column of value')
    col_id_attribute = fields.Integer(string='Column of Attributte ID')
    col_id_attribute_value = fields.Integer(string='Column of Attribute Value')

    def generate_random_name(self):
        letters = string.ascii_lowercase
        return ''.join(random.choice(letters) for i in range(10))

    def load_excel_file(self):
        if not self.col_attribute or not self.col_attribute_value:
            raise exceptions.Warning(_("The fields [attribute column] and [value column] are required"))
        else:
            if not self.serial_file:
                raise exceptions.Warning(_("You must select a file to modify"))
            else:

                try:
                    colattr = self.col_attribute  # columna 4 atributos
                    colattrv = self.col_attribute_value  # columna 5 valores de los atributos

                    final_name = "modified_file_" + self.generate_random_name()

                    with tempfile.NamedTemporaryFile(prefix=final_name, suffix='.xlsx', delete=False) as tmp:
                        tmp.write(base64.decodebytes(self.serial_file))

                    xfile = openpyxl.load_workbook(tmp.name)
                    sheet = xfile.worksheets[0]

                    row_count = sheet.max_row

                    row_with_value = 0

                    for row in range(1, row_count + 1):
                        if not sheet.cell(row, 1).value == None:
                            row_with_value = row
                        else:

                            if sheet.cell(row, colattr).value == None or sheet.cell(row,
                                                                                    colattr).value.strip() == sheet.cell(
                                row_with_value, colattr).value.strip():

                                sheet.cell(row_with_value, colattrv).value = sheet.cell(row_with_value,
                                                                                        colattrv).value.strip() + ',' + sheet.cell(
                                    row,
                                    colattrv).value.strip()
                                sheet.cell(row, colattrv).value = ""
                                sheet.cell(row, 1).value = 'delete'
                            else:
                                row_with_value = row

                    xfile.save(tmp.name)
                    self.clean_file_deleted_row(tmp)

                    with open(tmp.name, 'rb') as r:
                        file = base64.b64encode(r.read())

                    att_vals = {
                        'name': final_name + '.xlsx',
                        'type': 'binary',
                        'datas': file,
                    }

                    attachment_id = self.env['ir.attachment'].create(att_vals)
                    self.env.cr.commit()

                    return {
                        'type': 'ir.actions.act_url',
                        'url': '/web/content/{}?download=true'.format(attachment_id.id, ),
                        'target': 'self',
                    }
                except Exception:
                    raise exceptions.Warning(_("The selected file seems to be not valid"))

    def clean_file_deleted_row(self, tmp):
        xfile = openpyxl.load_workbook(tmp.name)
        sheet = xfile.worksheets[0]

        row_count = sheet.max_row
        column_count = sheet.max_column

        for row in range(1, row_count + 1):
            for row1 in range(row, row_count + 1):
                if sheet.cell(row1, 1).value == 'delete':
                    sheet.delete_rows(row1, 1)

        xfile.save(tmp.name)
